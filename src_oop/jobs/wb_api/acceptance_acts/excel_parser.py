"""Устойчивый Excel parser для актов приёма-передачи WB.

Слой читает один Excel-файл в памяти, определяет наиболее похожую структуру
таблицы и возвращает промежуточный результат `ExcelParseResult`. На этом этапе
здесь нет нормализации строк под БД, нет правил quantity/kiz и нет записи в БД.
"""

from __future__ import annotations

import io
import logging
import re
from collections.abc import Iterable
from datetime import date, datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from src_oop.jobs.wb_api.acceptance_acts.config import (
    ACT_TYPE_FBO,
    ACT_TYPE_FBS,
    CANONICAL_COLUMN_NAMES,
    EXCEL_COLUMN_SYNONYMS,
    FBO_SIGNATURE_FIELDS,
    FBS_SIGNATURE_FIELDS,
)
from src_oop.jobs.wb_api.acceptance_acts.models import (
    ExcelStructureInfo,
    ExcelParseResult,
    ExtractedExcelFile,
)

logger = logging.getLogger(__name__)

HEADER_SCAN_ROW_LIMIT = 25
HEADER_SCORE_MIN_THRESHOLD = 2.5
DATA_START_SCAN_LIMIT = 8
TOP_DATE_SCAN_ROWS = 8
TOP_DATE_SCAN_COLS = 8
KNOWN_DATE_CELLS = ("D3", "F3")
DATE_PATTERN = re.compile(r"(\d{1,2}[./-]\d{1,2}[./-]\d{2,4})")
PUNCTUATION_CLEAN_RE = re.compile(r"""["'`«».,:;]+""")
WHITESPACE_RE = re.compile(r"\s+")


class AcceptanceExcelParser:
    """Устойчивый parser Excel-файлов актов приёма-передачи WB.

    Слой поддерживает старую схему со строками 8/9/12 как один из кандидатов,
    но не завязывается на неё жёстко. Выбор структуры идёт через scoring по
    известным сигнатурам колонок и через fallback на соседние строки шапки.
    """

    def parse(self, excel_file: ExtractedExcelFile) -> ExcelParseResult:
        """Разбирает один Excel-файл и возвращает диагностический parse result.

        Алгоритм:
        1. Читает workbook и подбирает наиболее похожий рабочий лист.
        2. Находит дату документа.
        3. Подбирает одну или две строки шапки.
        4. Находит старт данных.
        5. Строит raw и canonical dataframe.
        6. Определяет фактический тип акта и итоговый статус.
        """
        result = ExcelParseResult(
            status="failed",
            account=excel_file.account,
            expected_act_type=excel_file.expected_act_type,
            actual_act_type=None,
            document_name=self._resolve_document_name(excel_file),
            document_number=excel_file.document_number_hint,
            document_date=None,
        )

        logger.info(
            "Старт parsing Excel WB: account=%s expected_act_type=%s excel=%s",
            excel_file.account,
            excel_file.expected_act_type,
            excel_file.excel_name,
        )

        try:
            workbook = self._load_workbook(excel_file, read_only=True)
        except Exception as error:
            message = f"Не удалось прочитать Excel-файл: {error}"
            result.errors.append(message)
            logger.exception(
                "Ошибка чтения Excel WB: account=%s excel=%s error=%s",
                excel_file.account,
                excel_file.excel_name,
                error,
            )
            return result

        worksheet, sheet_frame = self._select_working_sheet(workbook)
        if self._should_fallback_to_normal_mode(worksheet, sheet_frame):
            logger.warning(
                "Excel WB requires fallback to normal mode: account=%s excel=%s sheet=%s",
                excel_file.account,
                excel_file.excel_name,
                worksheet.title if worksheet is not None else None,
            )
            try:
                workbook = self._load_workbook(excel_file, read_only=False)
            except Exception as error:
                message = (
                    "Не удалось перечитать Excel-файл "
                    f"в fallback-режиме: {error}"
                )
                result.errors.append(message)
                logger.exception(
                    "Fallback-ошибка чтения Excel WB: account=%s excel=%s error=%s",
                    excel_file.account,
                    excel_file.excel_name,
                    error,
                )
                return result

            worksheet, sheet_frame = self._select_working_sheet(workbook)
            result.warnings.append(
                "Excel перечитан в normal mode из-за некорректного read_only sheet."
            )

        if worksheet is None or sheet_frame.empty:
            result.errors.append("Не удалось выбрать непустой рабочий лист Excel.")
            return result

        result.structure_info.sheet_name = worksheet.title

        document_date, selected_date_source, date_candidates = (
            self._detect_document_date(worksheet)
        )
        result.document_date = document_date
        result.structure_info.date_cell_candidates = date_candidates
        result.structure_info.selected_date_source = selected_date_source
        if document_date is None:
            result.warnings.append("Дата документа не найдена.")

        header_rows, structure_info = self._detect_header_rows(sheet_frame)
        result.structure_info.header_row_candidates = (
            structure_info.header_row_candidates
        )
        result.structure_info.selected_header_rows = (
            structure_info.selected_header_rows
        )
        result.structure_info.raw_headers = structure_info.raw_headers
        result.structure_info.canonical_headers = structure_info.canonical_headers
        result.structure_info.confidence_score = structure_info.confidence_score

        if not header_rows:
            result.errors.append("Не удалось распознать строки заголовков.")
            logger.warning(
                "Шапка Excel не распознана: account=%s excel=%s sheet=%s",
                excel_file.account,
                excel_file.excel_name,
                worksheet.title,
            )
            return result

        data_start_row = self._detect_data_start_row(sheet_frame, header_rows)
        result.structure_info.data_start_row = data_start_row
        if data_start_row is None:
            result.errors.append("Не удалось определить начало табличных данных.")
            return result

        raw_headers = self._build_combined_headers(sheet_frame, header_rows)
        normalized_headers = self._normalize_headers(raw_headers)
        header_mapping = self._map_headers_to_canonical_fields(normalized_headers)
        canonical_headers = [
            header_mapping.get(header, "") for header in normalized_headers
        ]

        result.structure_info.raw_headers = raw_headers
        result.structure_info.canonical_headers = canonical_headers

        raw_dataframe = self._build_raw_dataframe(
            sheet_frame=sheet_frame,
            raw_headers=raw_headers,
            data_start_row=data_start_row,
        )
        result.raw_dataframe = raw_dataframe

        if raw_dataframe.empty:
            result.errors.append("Табличные данные не найдены после шапки.")
            return result

        canonical_dataframe, unmapped_headers = self._build_canonical_dataframe(
            raw_dataframe=raw_dataframe,
            normalized_headers=normalized_headers,
            header_mapping=header_mapping,
            excel_file=excel_file,
            document_date=document_date,
            document_name=result.document_name,
        )
        result.canonical_dataframe = canonical_dataframe

        if unmapped_headers:
            result.warnings.append(
                "Не все заголовки сопоставлены с canonical names."
            )

        actual_act_type = self._detect_actual_act_type(
            canonical_dataframe.columns.tolist()
        )
        result.actual_act_type = actual_act_type
        if actual_act_type is None:
            result.warnings.append("Фактический тип акта определить не удалось.")
        else:
            missing_signature_fields = self._detect_missing_signature_fields(
                canonical_dataframe.columns.tolist(),
                actual_act_type=actual_act_type,
            )
            if missing_signature_fields:
                result.warnings.append(
                    "Часть сигнатурных колонок не найдена: "
                    + ", ".join(missing_signature_fields)
                )

        if structure_info.confidence_score < HEADER_SCORE_MIN_THRESHOLD + 0.5:
            result.warnings.append(
                "Использован fallback-подход для определения шапки таблицы."
            )

        result.status = self._resolve_status(result)

        logger.info(
            "Parsing Excel WB завершён: account=%s excel=%s sheet=%s header_rows=%s "
            "data_start=%s document_date=%s actual_act_type=%s status=%s warnings=%s errors=%s",
            excel_file.account,
            excel_file.excel_name,
            worksheet.title,
            result.structure_info.selected_header_rows,
            result.structure_info.data_start_row,
            result.document_date,
            result.actual_act_type,
            result.status,
            len(result.warnings),
            len(result.errors),
        )
        return result

    def _load_workbook(self, excel_file: ExtractedExcelFile, read_only: bool):
        """Читает workbook из `excel_bytes` без обращения к файловой системе."""
        return load_workbook(
            io.BytesIO(excel_file.excel_bytes),
            read_only=read_only,
            data_only=True,
        )

    def _select_working_sheet(
        self,
        workbook,
    ) -> tuple[Worksheet | None, pd.DataFrame]:
        """Выбирает наиболее содержательный лист для последующего parsing."""
        best_sheet: Worksheet | None = None
        best_frame = pd.DataFrame()
        best_score = -1

        for worksheet in workbook.worksheets:
            frame = self._worksheet_to_frame(worksheet)
            non_empty_cells = int(frame.notna().sum().sum()) if not frame.empty else 0
            if non_empty_cells > best_score:
                best_sheet = worksheet
                best_frame = frame
                best_score = non_empty_cells

        return best_sheet, best_frame

    def _detect_document_date(
        self,
        worksheet: Worksheet,
    ) -> tuple[date | None, str | None, list[str]]:
        """Ищет дату документа по known cells и по верхней области листа."""
        candidates: list[str] = []

        for cell_name in KNOWN_DATE_CELLS:
            value = worksheet[cell_name].value
            if value is None:
                continue
            candidates.append(f"{cell_name}={value}")
            parsed = self._parse_date_value(value)
            if parsed is not None:
                return parsed, cell_name, candidates

        for row in worksheet.iter_rows(
            min_row=1,
            max_row=TOP_DATE_SCAN_ROWS,
            min_col=1,
            max_col=TOP_DATE_SCAN_COLS,
            values_only=False,
        ):
            for cell in row:
                if cell.value in (None, ""):
                    continue
                candidates.append(f"{cell.coordinate}={cell.value}")
                parsed = self._parse_date_value(cell.value)
                if parsed is not None:
                    return parsed, cell.coordinate, candidates

        return None, None, candidates

    def _detect_header_rows(self, sheet_frame: pd.DataFrame) -> tuple[list[int], ExcelStructureInfo]:
        """Находит одну или две строки, наиболее похожие на шапку таблицы."""
        structure_info = ExcelStructureInfo()
        best_rows: list[int] = []
        best_score = -1.0
        best_raw_headers: list[str] = []
        best_canonical_headers: list[str] = []
        candidate_rows: list[int] = []

        scan_limit = min(len(sheet_frame), HEADER_SCAN_ROW_LIMIT)
        for row_index in range(scan_limit):
            candidate_rows.append(row_index)
            row_candidate = [row_index]
            score, raw_headers, canonical_headers = self._score_header_candidate(
                sheet_frame,
                row_candidate,
            )
            if score > best_score:
                best_rows = row_candidate
                best_score = score
                best_raw_headers = raw_headers
                best_canonical_headers = canonical_headers

            if row_index + 1 >= scan_limit:
                continue

            pair_candidate = [row_index, row_index + 1]
            score, raw_headers, canonical_headers = self._score_header_candidate(
                sheet_frame,
                pair_candidate,
            )
            if score > best_score:
                best_rows = pair_candidate
                best_score = score
                best_raw_headers = raw_headers
                best_canonical_headers = canonical_headers

        structure_info.header_row_candidates = candidate_rows
        structure_info.selected_header_rows = best_rows.copy()
        structure_info.raw_headers = best_raw_headers
        structure_info.canonical_headers = best_canonical_headers
        structure_info.confidence_score = max(best_score, 0.0)

        if best_score < HEADER_SCORE_MIN_THRESHOLD:
            return [], structure_info

        return best_rows, structure_info

    def _detect_data_start_row(
        self,
        sheet_frame: pd.DataFrame,
        header_rows: list[int],
    ) -> int | None:
        """Определяет первую строку данных после шапки."""
        if not header_rows:
            return None

        start_index = max(header_rows) + 1
        scan_end = min(len(sheet_frame), start_index + DATA_START_SCAN_LIMIT)

        for row_index in range(start_index, scan_end):
            row_values = sheet_frame.iloc[row_index].tolist()
            non_empty_count = self._count_non_empty_values(row_values)
            if non_empty_count >= 2 and not self._looks_like_helper_numbering_row(row_values):
                return row_index

        for row_index in range(start_index, len(sheet_frame)):
            row_values = sheet_frame.iloc[row_index].tolist()
            if (
                self._count_non_empty_values(row_values) > 0
                and not self._looks_like_helper_numbering_row(row_values)
            ):
                return row_index

        return None

    def _build_combined_headers(
        self,
        sheet_frame: pd.DataFrame,
        header_rows: list[int],
    ) -> list[str]:
        """Собирает заголовки из одной или двух строк, поддерживая merged-like форму."""
        if not header_rows:
            return []

        rows = [sheet_frame.iloc[row_index].tolist() for row_index in header_rows]
        max_columns = max(len(row) for row in rows)
        combined_headers: list[str] = []

        for column_index in range(max_columns):
            parts: list[str] = []
            for row in rows:
                value = row[column_index] if column_index < len(row) else None
                text = self._stringify_cell(value)
                if text:
                    parts.append(text)
            combined_headers.append(" - ".join(parts))

        return combined_headers

    def _normalize_headers(self, headers: list[str]) -> list[str]:
        """Нормализует сырые заголовки Excel перед mapping и scoring."""
        normalized_headers: list[str] = []
        for header in headers:
            value = self._stringify_cell(header).replace("\n", " ").replace("ё", "е")
            value = value.lower().strip()
            value = PUNCTUATION_CLEAN_RE.sub(" ", value)
            value = WHITESPACE_RE.sub(" ", value).strip()
            normalized_headers.append(value)
        return normalized_headers

    def _map_headers_to_canonical_fields(self, headers: list[str]) -> dict[str, str]:
        """Сопоставляет нормализованные заголовки каноническим полям WB."""
        synonyms_lookup = self._build_synonyms_lookup()
        mapping: dict[str, str] = {}

        for header in headers:
            if not header:
                continue
            canonical_name = synonyms_lookup.get(header)
            if canonical_name is not None:
                mapping[header] = canonical_name
                continue

            for synonym, candidate_name in synonyms_lookup.items():
                if synonym in header or header in synonym:
                    mapping[header] = candidate_name
                    break

        return mapping

    def _detect_actual_act_type(self, canonical_headers: list[str]) -> str | None:
        """Определяет фактический тип акта по каноническим колонкам."""
        header_set = {header for header in canonical_headers if header}

        has_fbs_signature = {
            "sticker",
            "order_number",
        }.issubset(header_set)
        has_fbo_signature = "box_barcode" in header_set and (
            "vendor_code" in header_set or "barcode" in header_set
        )

        if has_fbs_signature and not has_fbo_signature:
            return ACT_TYPE_FBS
        if has_fbo_signature and not has_fbs_signature:
            return ACT_TYPE_FBO
        return None

    def _worksheet_to_frame(self, worksheet: Worksheet) -> pd.DataFrame:
        """Преобразует worksheet в DataFrame без заголовков и без side effects."""
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return pd.DataFrame()
        return pd.DataFrame(rows, dtype=object)

    def _should_fallback_to_normal_mode(
        self,
        worksheet: Worksheet | None,
        sheet_frame: pd.DataFrame,
    ) -> bool:
        """Определяет, схлопнулся ли лист при чтении в `read_only=True`."""
        if worksheet is None:
            return True

        if sheet_frame.empty:
            return True

        if self._has_any_non_empty_cell(sheet_frame):
            return False

        max_row = getattr(worksheet, "max_row", None)
        max_column = getattr(worksheet, "max_column", None)
        if max_row is not None and max_column is not None and max_row <= 1 and max_column <= 1:
            return True

        try:
            if worksheet.calculate_dimension() == "A1:A1":
                return True
        except Exception:
            pass

        return self._top_area_looks_empty(worksheet)

    def _has_any_non_empty_cell(self, sheet_frame: pd.DataFrame) -> bool:
        """Проверяет, есть ли в DataFrame хотя бы одна непустая ячейка."""
        if sheet_frame.empty:
            return False

        for row in sheet_frame.itertuples(index=False, name=None):
            if self._count_non_empty_values(row) > 0:
                return True
        return False

    def _top_area_looks_empty(self, worksheet: Worksheet) -> bool:
        """Проверяет верхнюю область листа на тотальную пустоту."""
        for row in worksheet.iter_rows(
            min_row=1,
            max_row=TOP_DATE_SCAN_ROWS,
            min_col=1,
            max_col=TOP_DATE_SCAN_COLS,
            values_only=True,
        ):
            if self._count_non_empty_values(row) > 0:
                return False
        return True

    def _looks_like_helper_numbering_row(self, values: Iterable[object]) -> bool:
        """Определяет служебную строку нумерации колонок вида `1, 2, 3 ...`."""
        non_empty_values = [
            self._stringify_cell(value)
            for value in values
            if self._stringify_cell(value)
        ]
        if len(non_empty_values) < 2:
            return False

        numeric_values: list[int] = []
        for value in non_empty_values:
            if not value.isdigit():
                return False
            numeric_values.append(int(value))

        expected_sequence = list(range(1, len(numeric_values) + 1))
        return numeric_values == expected_sequence

    def _score_header_candidate(
        self,
        sheet_frame: pd.DataFrame,
        header_rows: list[int],
    ) -> tuple[float, list[str], list[str]]:
        """Оценивает кандидат на шапку по количеству распознанных колонок."""
        raw_headers = self._build_combined_headers(sheet_frame, header_rows)
        normalized_headers = self._normalize_headers(raw_headers)
        mapping = self._map_headers_to_canonical_fields(normalized_headers)
        canonical_headers = [mapping.get(header, "") for header in normalized_headers]

        unique_canonical = {header for header in canonical_headers if header}
        score = float(len(unique_canonical))

        if len(header_rows) == 2:
            score += 0.3
        if header_rows == [8, 9]:
            score += 0.7
        if header_rows == [8]:
            score += 0.3

        header_set = unique_canonical
        if set(FBS_SIGNATURE_FIELDS).issubset(header_set):
            score += 1.0
        if "box_barcode" in header_set and (
            "vendor_code" in header_set or "barcode" in header_set
        ):
            score += 1.0

        return score, raw_headers, canonical_headers

    def _build_canonical_dataframe(
        self,
        raw_dataframe: pd.DataFrame,
        normalized_headers: list[str],
        header_mapping: dict[str, str],
        excel_file: ExtractedExcelFile,
        document_date: date | None,
        document_name: str,
    ) -> tuple[pd.DataFrame, list[str]]:
        """Строит dataframe с canonical column names и тех. полями parser-а."""
        canonical_dataframe = pd.DataFrame(index=raw_dataframe.index.copy())
        unmapped_headers: list[str] = []

        for column_name, normalized_header in zip(
            raw_dataframe.columns.tolist(),
            normalized_headers,
            strict=False,
        ):
            canonical_name = header_mapping.get(normalized_header)
            if canonical_name is None:
                if normalized_header:
                    unmapped_headers.append(normalized_header)
                continue

            source_series = raw_dataframe[column_name]
            if canonical_name in canonical_dataframe.columns:
                canonical_dataframe[canonical_name] = canonical_dataframe[
                    canonical_name
                ].combine_first(source_series)
            else:
                canonical_dataframe[canonical_name] = source_series

        canonical_dataframe["document"] = document_name
        canonical_dataframe["account"] = excel_file.account
        if excel_file.document_number_hint is not None:
            canonical_dataframe["document_number"] = excel_file.document_number_hint
        if document_date is not None:
            canonical_dataframe["date"] = document_date

        ordered_columns = [
            column
            for column in CANONICAL_COLUMN_NAMES
            if column in canonical_dataframe.columns
        ]
        remaining_columns = [
            column
            for column in canonical_dataframe.columns
            if column not in ordered_columns
        ]
        canonical_dataframe = canonical_dataframe[ordered_columns + remaining_columns]

        return canonical_dataframe, unmapped_headers

    def _build_raw_dataframe(
        self,
        sheet_frame: pd.DataFrame,
        raw_headers: list[str],
        data_start_row: int,
    ) -> pd.DataFrame:
        """Строит сырую табличную часть с исходными заголовками."""
        header_count = len(raw_headers)
        data_frame = sheet_frame.iloc[data_start_row:, :header_count].copy()
        data_frame.columns = raw_headers
        data_frame = data_frame.dropna(how="all").reset_index(drop=True)
        non_empty_headers = [header if header else f"unnamed_{index}" for index, header in enumerate(data_frame.columns)]
        data_frame.columns = non_empty_headers
        return data_frame

    def _build_synonyms_lookup(self) -> dict[str, str]:
        """Строит lookup по нормализованным синонимам колонок WB."""
        lookup: dict[str, str] = {}
        for canonical_name, synonyms in EXCEL_COLUMN_SYNONYMS.items():
            lookup[self._normalize_headers([canonical_name])[0]] = canonical_name
            for synonym in synonyms:
                normalized = self._normalize_headers([synonym])[0]
                lookup[normalized] = canonical_name
        return lookup

    def _parse_date_value(self, value: object) -> date | None:
        """Пытается привести значение ячейки к дате документа."""
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value

        text = self._stringify_cell(value)
        if not text:
            return None

        cleaned = (
            text.replace('"', " ")
            .replace("г.", " ")
            .replace("г", " ")
            .replace(",", " ")
        )
        cleaned = WHITESPACE_RE.sub(" ", cleaned).strip()

        parsed = pd.to_datetime(cleaned, dayfirst=True, errors="coerce")
        if not pd.isna(parsed):
            return parsed.date()

        match = DATE_PATTERN.search(cleaned)
        if match:
            parsed = pd.to_datetime(match.group(1), dayfirst=True, errors="coerce")
            if not pd.isna(parsed):
                return parsed.date()

        return None

    def _resolve_document_name(self, excel_file: ExtractedExcelFile) -> str:
        """Выбирает наиболее стабильное имя документа для тех. поля `document`."""
        return excel_file.outer_entry_name or excel_file.excel_name

    def _resolve_status(self, result: ExcelParseResult) -> str:
        """Выставляет финальный статус parse result по результату разбора."""
        if result.errors:
            return "failed"
        if result.canonical_dataframe is None or result.canonical_dataframe.empty:
            return "failed"
        if result.actual_act_type is None or result.warnings:
            return "partial"
        return "success"

    def _count_non_empty_values(self, values: Iterable[object]) -> int:
        """Считает непустые значения в строке."""
        return sum(1 for value in values if self._stringify_cell(value))

    def _stringify_cell(self, value: object) -> str:
        """Безопасно приводит значение ячейки к строке для диагностики и matching."""
        if value is None:
            return ""
        if isinstance(value, float) and pd.isna(value):
            return ""
        text = str(value).strip()
        return "" if text.lower() == "nan" else text

    def _detect_missing_signature_fields(
        self,
        columns: list[str],
        actual_act_type: str,
    ) -> list[str]:
        """Возвращает сигнатурные поля, которые не встретились в canonical columns."""
        column_set = set(columns)
        if actual_act_type == ACT_TYPE_FBO:
            relevant_signatures = set(FBO_SIGNATURE_FIELDS)
        elif actual_act_type == ACT_TYPE_FBS:
            relevant_signatures = set(FBS_SIGNATURE_FIELDS)
        else:
            relevant_signatures = set()
        return sorted(field for field in relevant_signatures if field not in column_set)
