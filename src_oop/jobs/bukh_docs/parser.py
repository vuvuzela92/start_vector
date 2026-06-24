from __future__ import annotations

import base64
import io
import logging
import re
import zipfile
from collections.abc import Iterable

import pandas as pd
from openpyxl import load_workbook

from src_oop.jobs.bukh_docs.models import DownloadedDocumentsPayload, ExtractedFile

logger = logging.getLogger(__name__)

WEEKLY_REPORT_COLUMN_ALIASES = {
    "Наименование": "title",
    "Документ основание": "supporting_document",
    "Дата": "date",
    "№ документа": "doc_num",
    "Сумма, руб.": "sum_rub",
    "в т.ч НДС, руб.": "vat_rub",
}

REDEEM_NOTIFICATION_COLUMN_ALIASES = {
    "№": "№",
    "№ п/п": "№",
    "Артикул": "wild",
    "Наименование": "subject_name",
    "Количество": "quantity",
    "Сумма выкупа, руб., (вкл. НДС)": "sum_rub_with_vat",
    "Сумма выкупа, руб.": "sum_rub_with_vat",
    "Ставка НДС*, %": "vat_rate",
    "Ставка НДС": "vat_rate",
    "Сумма НДС*, Руб.": "vat_sum_rub",
    "Сумма НДС": "vat_sum_rub",
    "КИЗ": "kiz",
    "Документ": "doc_name",
}


class BukhDocsParser:
    """Парсер архивов и содержимого бухгалтерских документов WB."""

    def extract_files(
        self,
        payload: DownloadedDocumentsPayload,
    ) -> list[ExtractedFile]:
        doc_type_by_service_name = {
            document.service_name: document.doc_type
            for document in payload.document_requests
        }
        archive_bytes = base64.b64decode(payload.base64_document)
        extracted_files = self._extract_zip_files(
            account=payload.account,
            doc_type_by_service_name=doc_type_by_service_name,
            zip_bytes=archive_bytes,
        )
        logger.info(
            "Архив документов распакован: account=%s requested_documents=%s extracted_files=%s",
            payload.account,
            len(payload.document_requests),
            len(extracted_files),
        )
        return extracted_files

    def parse_weekly_reports(self, files: list[ExtractedFile]) -> pd.DataFrame:
        pdf_files = [file for file in files if file.path.lower().endswith(".pdf")]
        dataframes: list[pd.DataFrame] = []

        for file in pdf_files:
            table = self._extract_table_from_pdf(file.content)
            if not table:
                logger.warning(
                    "Не удалось извлечь таблицу из PDF weekly report: account=%s path=%s",
                    file.account,
                    file.path,
                )
                continue

            dataframe = pd.DataFrame(table[1:], columns=table[0])
            if dataframe.empty:
                logger.warning(
                    "PDF weekly report дал пустой DataFrame: account=%s path=%s",
                    file.account,
                    file.path,
                )
                continue

            dataframe.columns = [self._normalize_header(column) for column in dataframe.columns]
            if "Сумма, руб." in dataframe.columns:
                dataframe["Сумма, руб."] = dataframe["Сумма, руб."].map(self._clean_number)
            if "в т.ч НДС, руб." in dataframe.columns:
                dataframe["в т.ч НДС, руб."] = dataframe["в т.ч НДС, руб."].map(self._clean_number)
            if "Дата" in dataframe.columns:
                dataframe["Дата"] = pd.to_datetime(dataframe["Дата"], errors="coerce")

            dataframe["account"] = file.account.upper()
            dataframe = dataframe.rename(columns=WEEKLY_REPORT_COLUMN_ALIASES)
            dataframe = self._ensure_columns(
                dataframe=dataframe,
                required_columns=tuple(WEEKLY_REPORT_COLUMN_ALIASES.values()) + ("account",),
            )
            dataframe = dataframe.dropna(how="all")
            dataframes.append(dataframe)
            logger.info(
                "PDF weekly report распознан: account=%s path=%s rows=%s columns=%s",
                file.account,
                file.path,
                len(dataframe.index),
                list(dataframe.columns),
            )

        if not dataframes:
            return pd.DataFrame()

        result = pd.concat(dataframes, ignore_index=True)
        logger.info(
            "Собран итоговый weekly report DataFrame: files=%s rows=%s",
            len(dataframes),
            len(result.index),
        )
        return result

    def parse_redeem_notifications(self, files: list[ExtractedFile]) -> pd.DataFrame:
        xlsx_files = [file for file in files if file.path.lower().endswith(".xlsx")]
        dataframes: list[pd.DataFrame] = []

        for file in xlsx_files:
            try:
                dataframe = self._parse_single_redeem_file(file)
            except Exception as error:
                logger.warning(
                    "Ошибка обработки redeem Excel: account=%s path=%s error=%s",
                    file.account,
                    file.path,
                    error,
                )
                continue
            if dataframe.empty:
                logger.warning(
                    "Redeem Excel дал пустой DataFrame: account=%s path=%s",
                    file.account,
                    file.path,
                )
                continue
            dataframes.append(dataframe)
            logger.info(
                "Redeem Excel распознан: account=%s path=%s rows=%s columns=%s",
                file.account,
                file.path,
                len(dataframe.index),
                list(dataframe.columns),
            )

        if not dataframes:
            return pd.DataFrame()

        result = pd.concat(dataframes, ignore_index=True)
        result = result.loc[:, ~result.columns.isin([None, ""])]
        result.columns = [self._normalize_header(column) for column in result.columns]
        result = result.rename(columns=REDEEM_NOTIFICATION_COLUMN_ALIASES)
        result = self._ensure_columns(
            dataframe=result,
            required_columns=(
                "№",
                "wild",
                "subject_name",
                "quantity",
                "sum_rub_with_vat",
                "vat_rate",
                "vat_sum_rub",
                "kiz",
                "doc_name",
                "account",
            ),
        )
        if "wild" in result.columns:
            result["wild"] = result["wild"].astype(str).str.extract(r"(wild\d+)")
        logger.info(
            "Собран итоговый redeem notifications DataFrame: files=%s rows=%s",
            len(dataframes),
            len(result.index),
        )
        return result

    def _extract_zip_files(
        self,
        account: str,
        doc_type_by_service_name: dict[str, str],
        zip_bytes: bytes,
        base_path: str = "",
    ) -> list[ExtractedFile]:
        extracted_files: list[ExtractedFile] = []
        with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zip_file:
            for name in zip_file.namelist():
                full_path = f"{base_path}/{name}" if base_path else name
                data = zip_file.read(name)

                if name.lower().endswith(".zip") and data.startswith(b"PK"):
                    extracted_files.extend(
                        self._extract_zip_files(
                            account=account,
                            doc_type_by_service_name=doc_type_by_service_name,
                            zip_bytes=data,
                            base_path=full_path,
                        )
                    )
                    continue

                extracted_files.append(
                    ExtractedFile(
                        account=account,
                        doc_type=self._resolve_doc_type(full_path, doc_type_by_service_name),
                        path=full_path,
                        content=data,
                    )
                )
        return extracted_files

    def _resolve_doc_type(
        self,
        path: str,
        doc_type_by_service_name: dict[str, str],
    ) -> str:
        for service_name, doc_type in doc_type_by_service_name.items():
            if service_name in path:
                return doc_type
        return "unknown"

    def _extract_table_from_pdf(self, pdf_bytes: bytes) -> list[list[str]] | None:
        try:
            import pdfplumber
        except ModuleNotFoundError as error:
            raise ModuleNotFoundError(
                "Для обработки PDF-документов требуется пакет pdfplumber."
            ) from error

        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            if not pdf.pages:
                return None
            page = pdf.pages[0]
            table = page.extract_table()
            if table:
                return table
            tables = page.extract_tables()
            if tables:
                return tables[0]
        return None

    def _parse_single_redeem_file(self, file: ExtractedFile) -> pd.DataFrame:
        workbook = load_workbook(io.BytesIO(file.content), read_only=True)
        try:
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))

            table_start = None
            for index, row in enumerate(rows):
                if not row:
                    continue
                first_cell = self._normalize_header(row[0] if len(row) > 0 else None)
                second_cell = self._normalize_header(row[1] if len(row) > 1 else None)
                if first_cell and "№" in first_cell and (
                    "п/п" in first_cell or "Артикул" in second_cell
                ):
                    table_start = index
                    break

            if table_start is None:
                table_start = 9

            headers = [self._normalize_header(header) for header in rows[table_start]]
            data_rows = rows[table_start + 1 :]
            dataframe = pd.DataFrame(data_rows, columns=headers)
            if dataframe.empty:
                return dataframe

            first_column = dataframe.columns[0]
            dataframe = dataframe[dataframe[first_column] != "Итого:"]

            self._convert_excel_column_to_numeric(dataframe, 3, integer_mode=True)
            self._convert_excel_column_to_numeric(dataframe, 4)
            self._convert_excel_column_to_numeric(dataframe, 6)

            dataframe["Документ"] = sheet["A3"].value
            dataframe["account"] = file.account
            dataframe = dataframe[pd.to_numeric(dataframe.iloc[:, 0], errors="coerce").notna()]
            return dataframe
        finally:
            workbook.close()

    def _clean_number(self, value: object) -> float:
        if pd.isna(value) or value in ("—", "X", "", None):
            return 0.0

        normalized_value = str(value).replace("\xa0", " ")
        normalized_value = normalized_value.replace(" ", "").replace(",", ".")
        try:
            return float(normalized_value)
        except ValueError:
            return 0.0

    def _normalize_header(self, value: object) -> str:
        if value is None:
            return ""
        text_value = str(value).replace("\xa0", " ")
        text_value = re.sub(r"\s+", " ", text_value).strip()
        return text_value

    def _ensure_columns(
        self,
        dataframe: pd.DataFrame,
        required_columns: Iterable[str],
    ) -> pd.DataFrame:
        for column_name in required_columns:
            if column_name not in dataframe.columns:
                dataframe[column_name] = None
        return dataframe

    def _convert_excel_column_to_numeric(
        self,
        dataframe: pd.DataFrame,
        column_index: int,
        integer_mode: bool = False,
    ) -> None:
        if column_index >= len(dataframe.columns):
            return

        column_name = dataframe.columns[column_index]
        if integer_mode:
            dataframe[column_name] = pd.to_numeric(
                dataframe[column_name],
                errors="coerce",
            ).fillna(0)
            return

        dataframe[column_name] = dataframe[column_name].map(self._clean_number)
