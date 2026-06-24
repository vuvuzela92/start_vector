from __future__ import annotations

import io
import unittest
import zipfile
from openpyxl import Workbook

from src_oop.jobs.bukh_docs.models import (
    DocumentRequest,
    DownloadedDocumentsPayload,
    ExtractedFile,
)
from src_oop.jobs.bukh_docs.parser import BukhDocsParser


class BukhDocsParserTests(unittest.TestCase):
    def setUp(self) -> None:
        self.parser = BukhDocsParser()

    def test_extract_files_handles_nested_zip_and_doc_type_resolution(self) -> None:
        inner_zip_buffer = io.BytesIO()
        with zipfile.ZipFile(inner_zip_buffer, "w", zipfile.ZIP_DEFLATED) as inner_zip:
            inner_zip.writestr("weekly-implementation-report-1/report.pdf", b"fake-pdf")

        outer_zip_buffer = io.BytesIO()
        with zipfile.ZipFile(outer_zip_buffer, "w", zipfile.ZIP_DEFLATED) as outer_zip:
            outer_zip.writestr(
                "redeem-notification-1/archive.zip",
                inner_zip_buffer.getvalue(),
            )
            outer_zip.writestr("redeem-notification-1/redeem.xlsx", b"fake-xlsx")

        payload = DownloadedDocumentsPayload(
            account="acc1",
            document_requests=[
                DocumentRequest(
                    account="acc1",
                    doc_type="weekly-implementation-report",
                    service_name="weekly-implementation-report-1",
                    extension="pdf",
                ),
                DocumentRequest(
                    account="acc1",
                    doc_type="redeem-notification",
                    service_name="redeem-notification-1",
                    extension="xlsx",
                ),
            ],
            base64_document=self._to_base64(outer_zip_buffer.getvalue()),
        )

        extracted_files = self.parser.extract_files(payload)

        self.assertEqual(2, len(extracted_files))
        extracted_paths = {file.path: file.doc_type for file in extracted_files}
        self.assertEqual(
            "weekly-implementation-report",
            extracted_paths["redeem-notification-1/archive.zip/weekly-implementation-report-1/report.pdf"],
        )
        self.assertEqual(
            "redeem-notification",
            extracted_paths["redeem-notification-1/redeem.xlsx"],
        )

    def test_parse_weekly_reports_normalizes_columns_and_numbers(self) -> None:
        file = ExtractedFile(
            account="acc1",
            doc_type="weekly-implementation-report",
            path="doc.pdf",
            content=b"fake-pdf",
        )

        table = [
            [
                "Наименование",
                "Документ основание",
                "Дата",
                "№ документа",
                "Сумма, руб.",
                "в т.ч НДС, руб.",
            ],
            [
                "Отчет 1",
                "Основание 1",
                "2026-06-01",
                "101",
                "1 234,50",
                "234,50",
            ],
        ]

        self.parser._extract_table_from_pdf = lambda _: table  # type: ignore[method-assign]

        dataframe = self.parser.parse_weekly_reports([file])

        self.assertEqual(1, len(dataframe.index))
        self.assertEqual(
            [
                "title",
                "supporting_document",
                "date",
                "doc_num",
                "sum_rub",
                "vat_rub",
                "account",
            ],
            list(dataframe.columns),
        )
        self.assertEqual("Отчет 1", dataframe.iloc[0]["title"])
        self.assertEqual("ACC1", dataframe.iloc[0]["account"])
        self.assertEqual(1234.5, dataframe.iloc[0]["sum_rub"])
        self.assertEqual(234.5, dataframe.iloc[0]["vat_rub"])

    def test_parse_redeem_notifications_handles_generated_workbook(self) -> None:
        workbook_bytes = self._build_redeem_workbook_bytes()
        file = ExtractedFile(
            account="acc2",
            doc_type="redeem-notification",
            path="redeem.xlsx",
            content=workbook_bytes,
        )

        dataframe = self.parser.parse_redeem_notifications([file])

        self.assertEqual(1, len(dataframe.index))
        self.assertEqual("1", str(dataframe.iloc[0]["№"]))
        self.assertEqual("wild12345", dataframe.iloc[0]["wild"])
        self.assertEqual("Товар 1", dataframe.iloc[0]["subject_name"])
        self.assertEqual("Документ выкупа", dataframe.iloc[0]["doc_name"])
        self.assertEqual("acc2", dataframe.iloc[0]["account"])
        self.assertEqual(1500.5, dataframe.iloc[0]["sum_rub_with_vat"])
        self.assertEqual(250.5, dataframe.iloc[0]["vat_sum_rub"])

    def _build_redeem_workbook_bytes(self) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet["A3"] = "Документ выкупа"

        headers = [
            "№ \nп/п",
            "Артикул",
            "Наименование ",
            "Количество",
            "Сумма выкупа, руб., \n(вкл. НДС)",
            "Ставка НДС*,\n%",
            "Сумма НДС*,\nРуб.",
            "КИЗ",
        ]
        header_row_index = 10
        for column_index, value in enumerate(headers, start=1):
            sheet.cell(row=header_row_index, column=column_index, value=value)

        row_values = [
            1,
            "wild12345-extra",
            "Товар 1",
            2,
            "1 500,50",
            "20",
            "250,50",
            "KIZ-1",
        ]
        for column_index, value in enumerate(row_values, start=1):
            sheet.cell(row=header_row_index + 1, column=column_index, value=value)

        sheet.cell(row=header_row_index + 2, column=1, value="Итого:")

        buffer = io.BytesIO()
        workbook.save(buffer)
        workbook.close()
        return buffer.getvalue()

    def _to_base64(self, value: bytes) -> str:
        import base64

        return base64.b64encode(value).decode("ascii")


if __name__ == "__main__":
    unittest.main()
