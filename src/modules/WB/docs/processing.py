from src.modules.WB.docs.api import fetch_doc_list_wb, fetch_download_all_documents
from src.core.utils_general import load_api_tokens
import asyncio
import zipfile
import io
import base64
import pdfplumber
import pandas as pd
from openpyxl import load_workbook

async def processing_doc_list_wb(data=None)-> dict:
    """Функция получает и обрабатывает список доступных документов, возвращая словарь со списком документов, доступных для дальнейшего скачивания. 
    
    - data = await fetch_doc_list_wb(load_api_tokens()) возвращает список всех доступных для скачивания документов со всех аккаунтов

    """

    if data is None:
        data = await fetch_doc_list_wb(load_api_tokens())

    doc_dict = {}

    for first_level in data:
        if first_level:
            for doc in first_level:
                account = doc['account']
                
                if account not in doc_dict:
                    doc_dict[account] = []
                
                processed_doc = doc.copy()
                
                if processed_doc.get('extensions'):
                    processed_doc['extension'] = processed_doc['extensions'][0]
                
                delete_keys = ('name', 'category', 'creationTime', 'viewed', 'account', 'extensions')
                for key in delete_keys:
                    if key in processed_doc:
                        processed_doc.pop(key, None)  # безопасное удаление
                
                doc_dict[account].append(processed_doc)
                
    return doc_dict

def extract_all_files(zip_bytes, base_path=""):
    """Рекурсивно извлекает все файлы из ZIP (и вложенных ZIP) в словарь вида {имя_файла: байты}."""
    result = {}
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        for name in zf.namelist():
            full_name = f"{base_path}/{name}" if base_path else name
            with zf.open(name) as f:
                data = f.read()
                # Если это ZIP, рекурсивно заходим внутрь
                if name.lower().endswith('.zip') and data.startswith(b'PK'):
                    inner = extract_all_files(data, full_name)
                    result.update(inner)
                else:
                    result[full_name] = data
    return result


async def processed_zip_docs(docs_all=None):
    """Проходим по всем скачанным архивам.
    Функция extract_files_from_zip обрабатывает все архивированные файлы внутри скачанного архива. Возвращаем список словарей, где где ключ 'bytes' содержит байты конечных файлов (уже не архивов)"""
    if docs_all is None:
        docs_all = await fetch_download_all_documents()
    
    all_files = []  # список всех конечных файлов со всех документов
    for doc in docs_all:
        if doc:
            account = doc['account']
            doc_bytes = base64.b64decode(doc['data']['document'])
            files = extract_files_from_zip(doc_bytes)
            # Добавим информацию об аккаунте в каждый файл
            for f in files:
                f['account'] = account
            all_files.extend(files)
    return all_files

def extract_files_from_zip(zip_bytes, base_path=""):
    """
    Рекурсивно извлекает все конечные файлы из ZIP-архива.
    Возвращает список словарей: {'path': полный_путь, 'bytes': байты_файла}
    """
    result = []
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        for name in zf.namelist():
            full_path = f"{base_path}/{name}" if base_path else name
            # Читаем содержимое
            data = zf.read(name)
            # Если это ZIP, рекурсивно обрабатываем
            if name.lower().endswith('.zip') and data.startswith(b'PK'):
                # Рекурсивный вызов
                result.extend(extract_files_from_zip(data, full_path))
            else:
                # Конечный файл
                result.append({'path': full_path, 'bytes': data})
    return result


def extract_table_from_pdf(pdf_bytes):
    """
    Извлекает таблицу из PDF (первая страница).
    Возвращает список списков или None, если таблица не найдена.
    """
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        # Предполагаем, что таблица на первой странице
        page = pdf.pages[0]
        # Пробуем extract_table() — он ищет одну таблицу
        table = page.extract_table()
        if table:
            return table
        # Если не сработало, попробуем extract_tables() и возьмём первую
        tables = page.extract_tables()
        if tables:
            return tables[0]
    return None


# Функция для очистки числового значения
def clean_number(val):
    if pd.isna(val) or val in ('—', 'X', ''):
        return 0.0
    # Удаляем пробелы, заменяем запятую на точку
    val = str(val).replace(' ', '').replace(',', '.')
    try:
        return float(val)
    except ValueError:
        return 0.0

def processed_pdf_week_reps(pdf_files: list):
    """Функция обрабатывает pdf файлы, переданные списком в битовой формате и возвращает единый датафрейм. На вход ожидается список словарей с ключами bytes и account"""
    # Список для хранения извлеченных данных в виде таблиц
    pdf_tables = []
    for item in pdf_files:
        # Для каждого файла извлекаем битовые данные
        pdf_bytes = item['bytes']
        # Отдельно выносим аккаунт
        account = item['account']
        # Из pdf файла извлекаем таблицу
        table = extract_table_from_pdf(pdf_bytes)
        # Если таблица существует приводим ее к датафрейму для дальнейше обработки
        if table:
            df = pd.DataFrame(table[1:], columns=table[0])
            # Приводим числовые столбцы к float, убираем пробелы и заменяем запятую на точку
            if 'Сумма, руб.' in df.columns:
                df['Сумма, руб.'] = df['Сумма, руб.'].apply(clean_number)
            if 'в т.ч НДС, руб.' in df.columns:
                df['в т.ч НДС, руб.'] = df['в т.ч НДС, руб.'].apply(clean_number)
            if 'Дата' in df.columns:
                df['Дата'] = pd.to_datetime(df['Дата']).dt.date
            # Добавляем данные об аккаунте
            df['account'] = account.upper()
            df = df.rename(columns={"Наименование": "title",
                                    "Документ основание": "supporting_document",
                                    "Дата": "date",
                                    "№ документа": "doc_num",
                                    "Сумма, руб.": "sum_rub",
                                    "в т.ч НДС, руб.": "vat_rub"
                                    })
            pdf_tables.append(df)

    if pdf_tables:
        return pd.concat(pdf_tables, ignore_index=True)
    else:
        return pd.DataFrame() 
    

def process_redeem_file(item: dict):
    """
    Вспоомгательная функция.
    Обрабатывает один Excel-файл с уведомлением о выкупе.
    Возвращает DataFrame с данными таблицы и метаданными.
    """
    xlsx_bytes = item['bytes']
    account = item['account']
    
    # Загружаем workbook
    wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True)
    sheet = wb.active
  
    # Загружаем все строки в список
    rows = list(sheet.iter_rows(values_only=True))
    
    # Находим начало таблицы (где заголовки)
    table_start = None
    for i, row in enumerate(rows):
        if row[0] and '№' in str(row[0]) and ('п/п' in str(row[0]) or 'Артикул' in str(row[1])):
            table_start = i
            break
    
    if table_start is None:
        # Если не нашли, используем стандартную 9-ю строку (0-based)
        table_start = 9
    
    # Заголовки таблицы
    headers = rows[table_start]
    # Данные таблицы (начиная со следующей строки)
    data_rows = rows[table_start+1:]
    
    # Создаём DataFrame
    df = pd.DataFrame(data_rows, columns=headers)
    
    # Убираем строку с итогами (где в первой колонке 'Итого:')
    df = df[df.iloc[:, 0] != 'Итого:']
    
    # Преобразуем числовые колонки
    # Индексы колонок: 0 - №, 1 - Артикул, 2 - Наименование, 3 - Количество,
    # 4 - Сумма выкупа, 5 - Ставка НДС, 6 - Сумма НДС, 7 - КИЗ
    
    df.iloc[:, 3] = pd.to_numeric(df.iloc[:, 3], errors='coerce').fillna(0)
    df.iloc[:, 4] = df.iloc[:, 4].apply(clean_number)
    df.iloc[:, 6] = df.iloc[:, 6].apply(clean_number)
    
    # Добавляем метаданные
    doc_name = sheet['A3'].value
    df['Документ'] = doc_name
    df['account'] = account

    # Удаляем то, что снизу после таблицы в файле
    # 1. Находим индекс итоговой строки
    total_idx = None
    for i, val in enumerate(df.iloc[:, 0]):
        if val == 'Итого:':
            total_idx = i
            break

    # 2. Обрезаем DataFrame до итогов (если итоги найдены)
    if total_idx is not None:
        df = df.iloc[:total_idx]
    df = df[pd.to_numeric(df.iloc[:, 0], errors='coerce').notna()]
    
    return df


def process_xlsx_redeem_file(xlsx_files: dict):
    """
    Обрабатывает список Excel-файлов с уведомлениями о выкупе и возвращает объединённый DataFrame.
    
    Параметры:
    -----------
    xlsx_files : dict
        Список словарей, каждый из которых содержит:
        - 'bytes': байтовое содержимое Excel-файла
        - 'path': путь к файлу внутри архива
        - 'account': название аккаунта
        
    Возвращает:
    -----------
    pd.DataFrame
        Объединённый DataFrame со всеми записями из всех файлов,
        с корректными типами данных и без пустых колонок.
    """
    
    # Список для накопления DataFrame'ов из каждого файла
    all_reedems = []
    
    # Перебираем все переданные файлы
    for item in xlsx_files:
        try:
            # Обрабатываем отдельный файл с помощью вспомогательной функции
            df = process_redeem_file(item)
            # Добавляем результат в общий список
            all_reedems.append(df)
            print(f"✅ Обработан {item['path']}, строк: {len(df)}")
        except Exception as e:
            # Логируем ошибку, но не прерываем обработку остальных файлов
            print(f"❌ Ошибка в {item['path']}: {e}")

    # Если хотя бы один файл успешно обработан
    if all_reedems:
        # Объединяем все DataFrame'ы в один
        # ignore_index=True обеспечивает сквозную нумерацию строк
        df_reedems = pd.concat(all_reedems, ignore_index=True)

        # Удаляем колонки с пустыми именами
        # (например, если в исходном файле были лишние пустые столбцы)
        df_reedems = df_reedems.loc[:, ~df_reedems.columns.isin([None, ''])]

        # Приведение числовых колонок к типу float
        # Список колонок, которые должны содержать числа
        float_cols = ['Количество', 'Сумма выкупа, руб., \n(вкл. НДС)', 'Сумма НДС*,\nРуб.']
        
        for col in df_reedems.columns:
            # Если колонка есть в списке числовых - преобразуем её в float
            if col in float_cols:
                df_reedems[col] = df_reedems[col].astype(float)


        df_reedems = df_reedems.rename(columns={"№ \nп/п": "№",
                                                "Артикул": "wild",
                                                "Наименование ": "subject_name",
                                                "Количество": "quantity",
                                                "Сумма выкупа, руб., \n(вкл. НДС)": "sum_rub_with_vat",
                                                "Ставка НДС*,\n%": "vat_rate",
                                                "Сумма НДС*,\nРуб.": "vat_sum_rub",
                                                "КИЗ": "kiz",
                                                "Документ": "doc_name",
                                                })
        # Регулярное выражение для извлечения нужной части
        pattern = r'(wild\d+)'
        # Используем метод str.extract для извлечения нужного паттерна
        df_reedems['wild'] = df_reedems['wild'].str.extract(pattern)
        # Возвращаем итоговый DataFrame
        return df_reedems
    
    # Если ни один файл не обработан, возвращаем пустой DataFrame
    return pd.DataFrame()       

